"""
GBMS - Consulting Projects Routes
해외기술용역 프로젝트 관리 API
"""
from flask import Blueprint, request, jsonify, send_file
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import pandas as pd
from werkzeug.utils import secure_filename
from models import db, ConsultingProject, ActivityLog
from routes.auth import token_required
import os

consulting_bp = Blueprint('consulting', __name__)


def validate_project_data(data, is_update=False):
    """프로젝트 데이터 유효성 검증"""
    errors = []

    # 필수 필드 검증 (생성 시에만)
    if not is_update:
        if not data.get('titleKr'):
            errors.append('국문사업명은 필수 입력 항목입니다.')
        if not data.get('country'):
            errors.append('국가는 필수 입력 항목입니다.')

    # 수주년도 검증
    if 'contractYear' in data and data['contractYear']:
        try:
            year = int(data['contractYear'])
            if year < 1972 or year > 2100:
                errors.append('수주년도는 1972년부터 2100년 사이여야 합니다.')
        except (ValueError, TypeError):
            errors.append('수주년도는 숫자여야 합니다.')

    # 예산 검증
    if 'budget' in data and data['budget']:
        try:
            budget = float(data['budget'])
            if budget < 0:
                errors.append('예산은 0 이상이어야 합니다.')
        except (ValueError, TypeError):
            errors.append('예산은 숫자여야 합니다.')

    # 좌표 검증
    if 'latitude' in data and data['latitude']:
        try:
            lat = float(data['latitude'])
            if lat < -90 or lat > 90:
                errors.append('위도는 -90도에서 90도 사이여야 합니다.')
        except (ValueError, TypeError):
            errors.append('위도는 숫자여야 합니다.')

    if 'longitude' in data and data['longitude']:
        try:
            lon = float(data['longitude'])
            if lon < -180 or lon > 180:
                errors.append('경도는 -180도에서 180도 사이여야 합니다.')
        except (ValueError, TypeError):
            errors.append('경도는 숫자여야 합니다.')

    # 상태 검증
    if 'status' in data and data['status']:
        if data['status'] not in ['준공', '진행중']:
            errors.append('상태는 "준공" 또는 "진행중"이어야 합니다.')

    return errors


@consulting_bp.route('', methods=['GET'])
@token_required
def get_consulting_projects(current_user):
    """Get all consulting projects with filters"""
    # Get query parameters
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    country = request.args.get('country')
    status = request.args.get('status')
    year = request.args.get('year', type=int)
    client = request.args.get('client')
    search = request.args.get('search')

    # Build query
    query = ConsultingProject.query

    if country:
        query = query.filter(ConsultingProject.country == country)

    if status:
        query = query.filter(ConsultingProject.status == status)

    if year:
        query = query.filter(ConsultingProject.contract_year == year)

    if client:
        query = query.filter(ConsultingProject.client.ilike(f'%{client}%'))

    if search:
        query = query.filter(
            db.or_(
                ConsultingProject.title_kr.ilike(f'%{search}%'),
                ConsultingProject.title_en.ilike(f'%{search}%'),
                ConsultingProject.country.ilike(f'%{search}%'),
                ConsultingProject.client.ilike(f'%{search}%')
            )
        )

    # Order by contract year (descending) and number
    query = query.order_by(
        ConsultingProject.contract_year.desc(),
        ConsultingProject.number.asc()
    )

    # Paginate
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)

    return jsonify({
        'success': True,
        'data': [project.to_dict() for project in pagination.items],
        'total': pagination.total,
        'pages': pagination.pages,
        'currentPage': page
    })


@consulting_bp.route('/<int:project_id>', methods=['GET'])
@token_required
def get_consulting_project(current_user, project_id):
    """Get single consulting project"""
    project = ConsultingProject.query.get_or_404(project_id)

    return jsonify({
        'success': True,
        'data': project.to_dict()
    })


@consulting_bp.route('', methods=['POST'])
@token_required
def create_consulting_project(current_user):
    """Create new consulting project"""
    data = request.get_json()

    if not data:
        return jsonify({
            'success': False,
            'message': '요청 데이터가 없습니다.'
        }), 400

    # 데이터 유효성 검증
    errors = validate_project_data(data, is_update=False)
    if errors:
        return jsonify({
            'success': False,
            'message': '입력 데이터 검증 실패',
            'errors': errors
        }), 400

    # 중복 체크 (동일한 국가, 사업명, 수주년도)
    if data.get('titleKr') and data.get('country'):
        existing = ConsultingProject.query.filter_by(
            title_kr=data['titleKr'].strip(),
            country=data['country'].strip()
        )
        if data.get('contractYear'):
            existing = existing.filter_by(contract_year=data['contractYear'])

        if existing.first():
            return jsonify({
                'success': False,
                'message': '동일한 국가, 사업명, 수주년도를 가진 프로젝트가 이미 존재합니다.'
            }), 409

    try:
        # Create new project
        project = ConsultingProject(
            number=data.get('number'),
            contract_year=data.get('contractYear'),
            status=data.get('status', '준공'),
            country=data.get('country', '').strip(),
            latitude=data.get('latitude'),
            longitude=data.get('longitude'),
            title_en=data.get('titleEn', '').strip() if data.get('titleEn') else None,
            title_kr=data.get('titleKr', '').strip(),
            project_type=data.get('projectType', '').strip() if data.get('projectType') else None,
            start_date=data.get('startDate'),
            end_date=data.get('endDate'),
            budget=data.get('budget'),
            client=data.get('client', '').strip() if data.get('client') else None,
            created_by=current_user.id
        )

        db.session.add(project)
        db.session.flush()  # Get the project ID

        # Log activity
        log = ActivityLog(
            user_id=current_user.id,
            action='create',
            entity_type='consulting_project',
            entity_id=project.id,
            description=f'{current_user.name}님이 해외기술용역 프로젝트를 생성했습니다: {project.title_kr}',
            ip_address=request.remote_addr
        )
        db.session.add(log)

        db.session.commit()

        return jsonify({
            'success': True,
            'message': '프로젝트가 생성되었습니다.',
            'data': project.to_dict()
        }), 201

    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'프로젝트 생성 중 오류가 발생했습니다: {str(e)}'
        }), 500


@consulting_bp.route('/<int:project_id>', methods=['PUT'])
@token_required
def update_consulting_project(current_user, project_id):
    """Update consulting project"""
    project = ConsultingProject.query.get_or_404(project_id)
    data = request.get_json()

    if not data:
        return jsonify({
            'success': False,
            'message': '수정할 데이터가 없습니다.'
        }), 400

    # 데이터 유효성 검증
    errors = validate_project_data(data, is_update=True)
    if errors:
        return jsonify({
            'success': False,
            'message': '입력 데이터 검증 실패',
            'errors': errors
        }), 400

    # 중복 체크 (자기 자신은 제외)
    if data.get('titleKr') and data.get('country'):
        existing = ConsultingProject.query.filter(
            ConsultingProject.id != project_id,
            ConsultingProject.title_kr == data['titleKr'].strip(),
            ConsultingProject.country == data['country'].strip()
        )
        if data.get('contractYear'):
            existing = existing.filter_by(contract_year=data['contractYear'])

        if existing.first():
            return jsonify({
                'success': False,
                'message': '동일한 국가, 사업명, 수주년도를 가진 프로젝트가 이미 존재합니다.'
            }), 409

    try:
        # Update fields
        if 'number' in data:
            project.number = data['number']
        if 'contractYear' in data:
            project.contract_year = data['contractYear']
        if 'status' in data:
            project.status = data['status']
        if 'country' in data:
            project.country = data['country'].strip() if data['country'] else None
        if 'latitude' in data:
            project.latitude = data['latitude'] if data['latitude'] else None
        if 'longitude' in data:
            project.longitude = data['longitude'] if data['longitude'] else None
        if 'titleEn' in data:
            project.title_en = data['titleEn'].strip() if data['titleEn'] else None
        if 'titleKr' in data:
            project.title_kr = data['titleKr'].strip()
        if 'projectType' in data:
            project.project_type = data['projectType'].strip() if data['projectType'] else None
        if 'startDate' in data:
            project.start_date = data['startDate']
        if 'endDate' in data:
            project.end_date = data['endDate']
        if 'budget' in data:
            project.budget = data['budget'] if data['budget'] else None
        if 'client' in data:
            project.client = data['client'].strip() if data['client'] else None

        project.updated_at = datetime.utcnow()

        # Log activity
        log = ActivityLog(
            user_id=current_user.id,
            action='update',
            entity_type='consulting_project',
            entity_id=project.id,
            description=f'{current_user.name}님이 해외기술용역 프로젝트를 수정했습니다: {project.title_kr}',
            ip_address=request.remote_addr
        )
        db.session.add(log)

        db.session.commit()

        return jsonify({
            'success': True,
            'message': '프로젝트가 수정되었습니다.',
            'data': project.to_dict()
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'프로젝트 수정 중 오류가 발생했습니다: {str(e)}'
        }), 500


@consulting_bp.route('/<int:project_id>', methods=['DELETE'])
@token_required
def delete_consulting_project(current_user, project_id):
    """Delete consulting project"""
    project = ConsultingProject.query.get_or_404(project_id)

    project_title = project.title_kr

    # Log activity before deletion
    log = ActivityLog(
        user_id=current_user.id,
        action='delete',
        entity_type='consulting_project',
        entity_id=project.id,
        description=f'{current_user.name}님이 해외기술용역 프로젝트를 삭제했습니다: {project_title}',
        ip_address=request.remote_addr
    )
    db.session.add(log)

    db.session.delete(project)
    db.session.commit()

    return jsonify({
        'success': True,
        'message': '프로젝트가 삭제되었습니다.'
    })


@consulting_bp.route('/export', methods=['GET'])
@token_required
def export_consulting_projects(current_user):
    """Export consulting projects to Excel"""
    # Get filters from query parameters
    country = request.args.get('country')
    status = request.args.get('status')
    year = request.args.get('year', type=int)
    client = request.args.get('client')
    search = request.args.get('search')

    # Build query
    query = ConsultingProject.query

    if country:
        query = query.filter(ConsultingProject.country == country)
    if status:
        query = query.filter(ConsultingProject.status == status)
    if year:
        query = query.filter(ConsultingProject.contract_year == year)
    if client:
        query = query.filter(ConsultingProject.client.ilike(f'%{client}%'))
    if search:
        query = query.filter(
            db.or_(
                ConsultingProject.title_kr.ilike(f'%{search}%'),
                ConsultingProject.title_en.ilike(f'%{search}%'),
                ConsultingProject.country.ilike(f'%{search}%'),
                ConsultingProject.client.ilike(f'%{search}%')
            )
        )

    # Order by contract year and number
    query = query.order_by(
        ConsultingProject.contract_year.asc(),
        ConsultingProject.number.asc()
    )

    projects = query.all()

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "해외기술컨설팅"

    # Define headers
    headers = [
        '번호', '수주년도', '진행여부', '국가별', 'X', 'Y',
        '영문사업명', '국문사업명', '사업형태',
        '착수일', '준공일', '용역비(공사)(백만원)', '발주처'
    ]

    # Style for headers
    header_fill = PatternFill(start_color="0A3D62", end_color="0A3D62", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Write data
    for row_num, project in enumerate(projects, 2):
        ws.cell(row=row_num, column=1, value=project.number)
        ws.cell(row=row_num, column=2, value=project.contract_year)
        ws.cell(row=row_num, column=3, value=project.status)
        ws.cell(row=row_num, column=4, value=project.country)
        ws.cell(row=row_num, column=5, value=float(project.longitude) if project.longitude else None)
        ws.cell(row=row_num, column=6, value=float(project.latitude) if project.latitude else None)
        ws.cell(row=row_num, column=7, value=project.title_en)
        ws.cell(row=row_num, column=8, value=project.title_kr)
        ws.cell(row=row_num, column=9, value=project.project_type)
        ws.cell(row=row_num, column=10, value=project.start_date)
        ws.cell(row=row_num, column=11, value=project.end_date)
        ws.cell(row=row_num, column=12, value=float(project.budget) if project.budget else None)
        ws.cell(row=row_num, column=13, value=project.client)

    # Adjust column widths
    column_widths = [8, 10, 10, 15, 12, 12, 35, 35, 20, 12, 12, 15, 20]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Log activity
    log = ActivityLog(
        user_id=current_user.id,
        action='export',
        entity_type='consulting_project',
        description=f'{current_user.name}님이 해외기술용역 프로젝트 {len(projects)}건을 Excel로 다운로드했습니다.',
        ip_address=request.remote_addr
    )
    db.session.add(log)
    db.session.commit()

    # Generate filename with timestamp
    filename = f"해외기술용역_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@consulting_bp.route('/stats', methods=['GET'])
@token_required
def get_consulting_stats(current_user):
    """Get consulting projects statistics"""
    # Total projects
    total = ConsultingProject.query.count()

    # By status
    status_stats = db.session.query(
        ConsultingProject.status,
        db.func.count(ConsultingProject.id)
    ).group_by(ConsultingProject.status).all()

    # By country (top 10)
    country_stats = db.session.query(
        ConsultingProject.country,
        db.func.count(ConsultingProject.id)
    ).group_by(ConsultingProject.country).order_by(
        db.func.count(ConsultingProject.id).desc()
    ).limit(10).all()

    # By year (last 10 years)
    year_stats = db.session.query(
        ConsultingProject.contract_year,
        db.func.count(ConsultingProject.id),
        db.func.sum(ConsultingProject.budget)
    ).group_by(ConsultingProject.contract_year).order_by(
        ConsultingProject.contract_year.desc()
    ).limit(10).all()

    # Total budget
    total_budget = db.session.query(
        db.func.sum(ConsultingProject.budget)
    ).scalar() or 0

    # Projects with coordinates
    with_coords = ConsultingProject.query.filter(
        ConsultingProject.latitude.isnot(None),
        ConsultingProject.longitude.isnot(None)
    ).count()

    return jsonify({
        'success': True,
        'data': {
            'total': total,
            'byStatus': {status: count for status, count in status_stats},
            'byCountry': [{'country': country, 'count': count} for country, count in country_stats],
            'byYear': [
                {
                    'year': year,
                    'count': count,
                    'budget': float(budget) if budget else 0
                }
                for year, count, budget in year_stats
            ],
            'totalBudget': float(total_budget),
            'withCoordinates': with_coords,
            'coordinateRate': round(with_coords / total * 100, 1) if total > 0 else 0
        }
    })


@consulting_bp.route('/countries', methods=['GET'])
@token_required
def get_consulting_countries(current_user):
    """Get list of countries with projects"""
    countries = db.session.query(
        ConsultingProject.country,
        db.func.count(ConsultingProject.id).label('count')
    ).group_by(ConsultingProject.country).order_by(
        ConsultingProject.country.asc()
    ).all()

    return jsonify({
        'success': True,
        'data': [
            {'country': country, 'count': count}
            for country, count in countries
        ]
    })


@consulting_bp.route('/clients', methods=['GET'])
@token_required
def get_consulting_clients(current_user):
    """Get list of clients"""
    clients = db.session.query(
        ConsultingProject.client
    ).filter(
        ConsultingProject.client.isnot(None)
    ).distinct().order_by(
        ConsultingProject.client.asc()
    ).all()

    return jsonify({
        'success': True,
        'data': [client[0] for client in clients if client[0]]
    })


@consulting_bp.route('/upload', methods=['POST'])
@token_required
def upload_consulting_projects(current_user):
    """Excel 파일을 통한 프로젝트 일괄 업로드"""

    if 'file' not in request.files:
        return jsonify({
            'success': False,
            'message': '파일이 전송되지 않았습니다.'
        }), 400

    file = request.files['file']

    if file.filename == '':
        return jsonify({
            'success': False,
            'message': '파일이 선택되지 않았습니다.'
        }), 400

    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({
            'success': False,
            'message': 'Excel 파일(.xlsx, .xls)만 업로드 가능합니다.'
        }), 400

    try:
        # Excel 파일 읽기
        df = pd.read_excel(file)

        # 필수 컬럼 확인
        required_columns = ['국문사업명', '국가별']
        missing_columns = [col for col in required_columns if col not in df.columns]

        if missing_columns:
            return jsonify({
                'success': False,
                'message': f'필수 컬럼이 없습니다: {", ".join(missing_columns)}'
            }), 400

        # 컬럼 매핑
        column_mapping = {
            '번호': 'number',
            '수주년도': 'contract_year',
            '진행여부': 'status',
            '국가별': 'country',
            'X': 'longitude',
            'Y': 'latitude',
            '영문사업명': 'title_en',
            '국문사업명': 'title_kr',
            '사업형태': 'project_type',
            '착수일': 'start_date',
            '준공일': 'end_date',
            '용역비(공사)(백만원)': 'budget',
            '발주처': 'client'
        }

        imported_count = 0
        skipped_count = 0
        errors = []

        for idx, row in df.iterrows():
            try:
                # 필수 필드 검증
                if pd.isna(row.get('국문사업명')) or pd.isna(row.get('국가별')):
                    errors.append(f'행 {idx + 2}: 필수 필드(국문사업명, 국가) 누락')
                    skipped_count += 1
                    continue

                # 프로젝트 데이터 준비
                project_data = {
                    'number': int(row['번호']) if pd.notna(row.get('번호')) else None,
                    'contract_year': int(row['수주년도']) if pd.notna(row.get('수주년도')) else None,
                    'status': row['진행여부'] if pd.notna(row.get('진행여부')) else '준공',
                    'country': str(row['국가별']).strip(),
                    'longitude': float(row['X']) if pd.notna(row.get('X')) else None,
                    'latitude': float(row['Y']) if pd.notna(row.get('Y')) else None,
                    'title_en': str(row['영문사업명']).strip() if pd.notna(row.get('영문사업명')) else None,
                    'title_kr': str(row['국문사업명']).strip(),
                    'project_type': str(row['사업형태']).strip() if pd.notna(row.get('사업형태')) else None,
                    'start_date': str(row['착수일']) if pd.notna(row.get('착수일')) else None,
                    'end_date': str(row['준공일']) if pd.notna(row.get('준공일')) else None,
                    'budget': float(row['용역비(공사)(백만원)']) if pd.notna(row.get('용역비(공사)(백만원)')) else None,
                    'client': str(row['발주처']).strip() if pd.notna(row.get('발주처')) else None,
                    'created_by': current_user.id
                }

                # 중복 체크
                existing = ConsultingProject.query.filter_by(
                    title_kr=project_data['title_kr'],
                    country=project_data['country']
                )
                if project_data['contract_year']:
                    existing = existing.filter_by(contract_year=project_data['contract_year'])

                if existing.first():
                    errors.append(f'행 {idx + 2}: 중복된 프로젝트 - {project_data["title_kr"]}')
                    skipped_count += 1
                    continue

                # 프로젝트 생성
                project = ConsultingProject(**project_data)
                db.session.add(project)
                imported_count += 1

            except Exception as e:
                errors.append(f'행 {idx + 2}: {str(e)}')
                skipped_count += 1
                continue

        # 데이터베이스에 커밋
        db.session.commit()

        # 활동 로그
        log = ActivityLog(
            user_id=current_user.id,
            action='import',
            entity_type='consulting_project',
            description=f'{current_user.name}님이 Excel 파일로 {imported_count}개의 해외기술용역 프로젝트를 업로드했습니다.',
            ip_address=request.remote_addr
        )
        db.session.add(log)
        db.session.commit()

        return jsonify({
            'success': True,
            'message': f'업로드가 완료되었습니다. (성공: {imported_count}개, 실패: {skipped_count}개)',
            'data': {
                'imported': imported_count,
                'skipped': skipped_count,
                'total': imported_count + skipped_count,
                'errors': errors[:10]  # 최대 10개의 에러만 반환
            }
        }), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'업로드 중 오류가 발생했습니다: {str(e)}'
        }), 500


@consulting_bp.route('/bulk-delete', methods=['POST'])
@token_required
def bulk_delete_consulting_projects(current_user):
    """여러 프로젝트를 일괄 삭제"""
    data = request.get_json()

    if not data or 'ids' not in data:
        return jsonify({
            'success': False,
            'message': '삭제할 프로젝트 ID 목록이 없습니다.'
        }), 400

    ids = data['ids']

    if not isinstance(ids, list) or len(ids) == 0:
        return jsonify({
            'success': False,
            'message': '유효한 ID 목록이 아닙니다.'
        }), 400

    try:
        # 프로젝트 조회
        projects = ConsultingProject.query.filter(ConsultingProject.id.in_(ids)).all()

        if not projects:
            return jsonify({
                'success': False,
                'message': '삭제할 프로젝트를 찾을 수 없습니다.'
            }), 404

        deleted_count = len(projects)
        project_titles = [p.title_kr for p in projects[:5]]  # 최대 5개만 기록

        # 프로젝트 삭제
        for project in projects:
            db.session.delete(project)

        # 활동 로그
        log = ActivityLog(
            user_id=current_user.id,
            action='bulk_delete',
            entity_type='consulting_project',
            description=f'{current_user.name}님이 {deleted_count}개의 해외기술용역 프로젝트를 일괄 삭제했습니다.',
            ip_address=request.remote_addr
        )
        db.session.add(log)

        db.session.commit()

        return jsonify({
            'success': True,
            'message': f'{deleted_count}개의 프로젝트가 삭제되었습니다.',
            'data': {
                'deleted': deleted_count,
                'titles': project_titles
            }
        }), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'삭제 중 오류가 발생했습니다: {str(e)}'
        }), 500
